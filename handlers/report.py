"""
Report Handler - Báo cáo tồn kho, xuất nhập, lãi lỗ, xuất Excel
"""
import io
import os
from datetime import datetime, timedelta
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext
from sqlalchemy.ext.asyncio import AsyncSession

from database.crud import (
    get_all_materials, get_transactions, get_low_stock_materials,
    calculate_profit_by_material, calculate_period_summary,
)
from keyboards.inline import (
    report_menu_keyboard, period_keyboard, main_menu_keyboard, back_keyboard,
)
from utils.formatters import (
    format_currency, format_number, format_date, format_datetime,
    format_profit_indicator,
)

router = Router()


def get_period_dates(period: str) -> tuple[datetime | None, datetime | None]:
    """Tính start_date và end_date theo kỳ"""
    now = datetime.now()
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now
    elif period == "7d":
        return now - timedelta(days=7), now
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    elif period == "quarter":
        quarter_month = ((now.month - 1) // 3) * 3 + 1
        start = now.replace(month=quarter_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    elif period == "year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    else:  # all
        return None, None


# ============================================================
# MENU BÁO CÁO
# ============================================================
@router.callback_query(F.data == "menu_report")
async def show_report_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "📊 <b>BÁO CÁO</b>\n\nChọn loại báo cáo:",
        reply_markup=report_menu_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data == "menu_profit")
async def show_profit_menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "💰 <b>LÃI / LỖ</b>\n\nChọn loại báo cáo:",
        reply_markup=report_menu_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


# ============================================================
# BÁO CÁO TỒN KHO
# ============================================================
@router.callback_query(F.data == "report_stock")
async def report_stock(callback: CallbackQuery, session: AsyncSession):
    materials = await get_all_materials(session)

    if not materials:
        await callback.message.edit_text(
            "📦 Chưa có vật tư nào.",
            reply_markup=report_menu_keyboard(),
            parse_mode="HTML",
        )
        await callback.answer()
        return

    now = format_date(datetime.now())
    text = f"📦 <b>BÁO CÁO TỒN KHO</b>\n📅 {now}\n\n"

    total_value = 0
    for m in materials:
        value = m.current_stock * m.cost_price
        total_value += value
        icon = "🟢" if m.current_stock >= m.min_stock or m.min_stock == 0 else "🔴"
        text += (
            f"{icon} <code>{m.material_code}</code>\n"
            f"   {m.name}: <b>{format_number(m.current_stock)}</b> {m.unit}\n"
            f"   Giá trị: {format_currency(value)}\n\n"
        )

    text += (
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 Tổng loại vật tư: <b>{len(materials)}</b>\n"
        f"💰 Tổng giá trị tồn kho: <b>{format_currency(total_value)}</b>"
    )

    await callback.message.edit_text(
        text, reply_markup=report_menu_keyboard(), parse_mode="HTML",
    )
    await callback.answer()


# ============================================================
# BÁO CÁO NHẬP/XUẤT KHO
# ============================================================
@router.callback_query(F.data == "report_import")
async def report_import_select_period(callback: CallbackQuery, state: FSMContext):
    await state.update_data(report_type="IMPORT")
    await callback.message.edit_text(
        "📥 <b>BÁO CÁO NHẬP KHO</b>\n\nChọn kỳ:",
        reply_markup=period_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data == "report_export")
async def report_export_select_period(callback: CallbackQuery, state: FSMContext):
    await state.update_data(report_type="EXPORT")
    await callback.message.edit_text(
        "📤 <b>BÁO CÁO XUẤT KHO</b>\n\nChọn kỳ:",
        reply_markup=period_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data.startswith("period_"))
async def report_by_period(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    period = callback.data.replace("period_", "")
    start_date, end_date = get_period_dates(period)
    data = await state.get_data()
    report_type = data.get("report_type")

    if report_type in ("IMPORT", "EXPORT"):
        # Báo cáo nhập/xuất kho
        transactions = await get_transactions(session, trans_type=report_type,
                                               start_date=start_date, end_date=end_date)
        icon = "📥" if report_type == "IMPORT" else "📤"
        title = "NHẬP KHO" if report_type == "IMPORT" else "XUẤT KHO"

        period_text = f"{format_date(start_date)} - {format_date(end_date)}" if start_date else "Tất cả"
        text = f"{icon} <b>BÁO CÁO {title}</b>\n📅 {period_text}\n\n"

        if not transactions:
            text += "Không có giao dịch nào trong kỳ."
        else:
            total = 0
            for t in transactions[:30]:
                material = await session.get(
                    __import__('database.models', fromlist=['Material']).Material,
                    t.material_id
                )
                mat_name = material.name if material else "?"
                text += (
                    f"📋 {format_datetime(t.created_at)}\n"
                    f"   {mat_name}: {format_number(t.quantity)} × {format_currency(t.unit_price)}\n"
                    f"   = {format_currency(t.total_amount)}\n\n"
                )
                total += t.total_amount

            text += (
                "━━━━━━━━━━━━━━━━━━━━\n"
                f"📊 Tổng GD: <b>{len(transactions)}</b>\n"
                f"💰 Tổng tiền: <b>{format_currency(total)}</b>"
            )

        await callback.message.edit_text(
            text, reply_markup=report_menu_keyboard(), parse_mode="HTML",
        )

    elif report_type == "profit_product":
        # Lãi lỗ theo sản phẩm
        results = await calculate_profit_by_material(session, start_date, end_date)
        period_text = f"{format_date(start_date)} - {format_date(end_date)}" if start_date else "Tất cả"
        text = f"💰 <b>LÃI LỖ THEO SẢN PHẨM</b>\n📅 {period_text}\n\n"

        if not results:
            text += "Chưa có giao dịch xuất kho nào."
        else:
            total_revenue = 0
            total_cogs = 0
            total_profit = 0
            loss_count = 0

            for r in results:
                m = r["material"]
                icon = "📈" if r["profit"] >= 0 else "📉"
                text += (
                    f"{icon} <b>{m.name}</b>\n"
                    f"   Doanh thu: {format_currency(r['revenue'])}\n"
                    f"   Giá vốn:  {format_currency(r['cogs'])}\n"
                    f"   Lãi/Lỗ:   <b>{format_currency(r['profit'])}</b> ({r['margin']:.1f}%)\n\n"
                )
                total_revenue += r["revenue"]
                total_cogs += r["cogs"]
                total_profit += r["profit"]
                if r["profit"] < 0:
                    loss_count += 1

            text += (
                "━━━━━━━━━━━━━━━━━━━━\n"
                f"💵 Tổng doanh thu: <b>{format_currency(total_revenue)}</b>\n"
                f"🏷️ Tổng giá vốn: {format_currency(total_cogs)}\n"
                f"{'📈' if total_profit >= 0 else '📉'} Tổng lãi/lỗ: <b>{format_currency(total_profit)}</b>\n"
            )
            if loss_count > 0:
                text += f"⚠️ SP lỗ: {loss_count}/{len(results)}"

        await callback.message.edit_text(
            text, reply_markup=report_menu_keyboard(), parse_mode="HTML",
        )

    elif report_type == "profit_period":
        # Tổng hợp lãi lỗ theo kỳ
        summary = await calculate_period_summary(session, start_date, end_date)
        period_text = f"{format_date(start_date)} - {format_date(end_date)}" if start_date else "Tất cả"

        profit_icon = "📈" if summary["gross_profit"] >= 0 else "📉"
        text = (
            f"💰 <b>TỔNG HỢP LÃI LỖ</b>\n📅 {period_text}\n\n"
            f"📥 Tổng nhập kho: <b>{format_currency(summary['total_import'])}</b>\n"
            f"   ({summary['import_count']} phiếu nhập)\n\n"
            f"📤 Tổng xuất kho: <b>{format_currency(summary['total_export'])}</b>\n"
            f"   ({summary['export_count']} phiếu xuất)\n\n"
            f"🏷️ Giá vốn hàng xuất: {format_currency(summary['total_cogs'])}\n\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"{profit_icon} <b>LÃI GỘP: {format_currency(summary['gross_profit'])}</b>\n"
            f"📊 Biên lợi nhuận: <b>{summary['margin']:.1f}%</b>"
        )

        await callback.message.edit_text(
            text, reply_markup=report_menu_keyboard(), parse_mode="HTML",
        )

    await state.clear()
    await callback.answer()


# ============================================================
# LÃI LỖ THEO SẢN PHẨM / THEO KỲ
# ============================================================
@router.callback_query(F.data == "report_profit_product")
async def profit_by_product(callback: CallbackQuery, state: FSMContext):
    await state.update_data(report_type="profit_product")
    await callback.message.edit_text(
        "💰 <b>LÃI LỖ THEO SẢN PHẨM</b>\n\nChọn kỳ:",
        reply_markup=period_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data == "report_profit_period")
async def profit_by_period(callback: CallbackQuery, state: FSMContext):
    await state.update_data(report_type="profit_period")
    await callback.message.edit_text(
        "📈 <b>TỔNG HỢP LÃI LỖ THEO KỲ</b>\n\nChọn kỳ:",
        reply_markup=period_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


# ============================================================
# XUẤT EXCEL
# ============================================================
@router.callback_query(F.data == "report_excel")
async def export_excel(callback: CallbackQuery, session: AsyncSession):
    materials = await get_all_materials(session)

    if not materials:
        await callback.message.edit_text(
            "📦 Chưa có dữ liệu để xuất.",
            reply_markup=report_menu_keyboard(),
            parse_mode="HTML",
        )
        await callback.answer()
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Tồn Kho"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"BÁO CÁO TỒN KHO - {format_date(datetime.now())}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["STT", "Mã vật tư", "Tên vật tư", "ĐVT", "Giá vốn", "Giá bán", "Tồn kho", "Giá trị tồn"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        total_value = 0
        for i, m in enumerate(materials, 1):
            value = m.current_stock * m.cost_price
            total_value += value
            row = i + 3
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=m.material_code).border = thin_border
            ws.cell(row=row, column=3, value=m.name).border = thin_border
            ws.cell(row=row, column=4, value=m.unit).border = thin_border
            ws.cell(row=row, column=5, value=m.cost_price).border = thin_border
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6, value=m.selling_price).border = thin_border
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=7, value=m.current_stock).border = thin_border
            ws.cell(row=row, column=7).number_format = '#,##0.#'
            ws.cell(row=row, column=8, value=value).border = thin_border
            ws.cell(row=row, column=8).number_format = '#,##0'

        # Total row
        total_row = len(materials) + 4
        ws.cell(row=total_row, column=7, value="TỔNG:").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_value).font = Font(bold=True)
        ws.cell(row=total_row, column=8).number_format = '#,##0'

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"TonKho_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        file = BufferedInputFile(buffer.read(), filename=filename)

        await callback.message.answer_document(
            file,
            caption=f"📎 <b>Báo cáo tồn kho</b>\n📅 {format_date(datetime.now())}\n"
                    f"📊 {len(materials)} loại vật tư\n"
                    f"💰 Tổng giá trị: {format_currency(total_value)}",
            parse_mode="HTML",
        )
        await callback.message.edit_text(
            "✅ Đã xuất file Excel!",
            reply_markup=report_menu_keyboard(),
            parse_mode="HTML",
        )
    except Exception as e:
        await callback.message.edit_text(
            f"❌ Lỗi xuất Excel: {str(e)}",
            reply_markup=report_menu_keyboard(),
            parse_mode="HTML",
        )
    await callback.answer()


# ============================================================
# CÀI ĐẶT
# ============================================================
@router.callback_query(F.data == "menu_settings")
async def show_settings(callback: CallbackQuery, session: AsyncSession):
    from database.crud import get_user_by_telegram_id
    user = await get_user_by_telegram_id(session, callback.from_user.id)

    role_text = {
        "ADMIN": "👑 Quản trị viên",
        "ACCOUNTANT": "📋 Kế toán",
        "VIEWER": "👁️ Xem",
    }

    text = (
        "⚙️ <b>CÀI ĐẶT</b>\n\n"
        f"👤 Tên: <b>{callback.from_user.full_name}</b>\n"
        f"🆔 Telegram ID: <code>{callback.from_user.id}</code>\n"
        f"🔑 Quyền: {role_text.get(user.role, user.role) if user else 'Chưa đăng ký'}\n\n"
        "<i>💡 Gửi Telegram ID cho Admin để được phân quyền cao hơn</i>"
    )

    await callback.message.edit_text(
        text, reply_markup=main_menu_keyboard(), parse_mode="HTML",
    )
    await callback.answer()
